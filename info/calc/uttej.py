from django.shortcuts import render
from django.http import HttpResponse
import pandas as pd
import openpyxl
import requests
#import pygmaps
import statistics
import gmplot
import numpy as np
import pandas as pd
import pyDEA
from scipy.optimize import fmin_slsqp
from io import BytesIO

def check_extension(name):
    if name.endswith('.xlsx'):
        return True
    else:
        return False

# Create your views here.

#def home(request):
 #   return render(request,'ms.html')
def mapon(request):
    return render(request,'new3.html')



def home_performance(request):
    if request.method=="POST":
      excel_file = request.FILES["myfile"]
      check=check_extension(excel_file.name)
      if check is False:
          return render(request,'PVS.html',context={"error_in_file":"Allowed file extension is xlsx."})

      workbook = openpyxl.load_workbook(excel_file, data_only=True)
      worksheet=workbook.active
      excel_data = list()
      for row in worksheet.iter_rows():
          row_data = list()
          for cell in row:
              if cell.value=='':
                  cell.value=None
              row_data.append(cell.value)
          excel_data.append(row_data)
      upload_list=excel_data[0]
      valid_data=excel_data[1:]

      df = pd.DataFrame(columns=excel_data[0], data=excel_data[1:])
      print(df)
      upload_list=df.columns.tolist()
      check_list=['place','performance value']
      if check_list!=upload_list:
          return render(request,'PVS.html',{'name':'uttej','error_in_file':'Headers not matched.Alllowed headers is place,performance value !'})
      else:
          url = "https://maps.googleapis.com/maps/api/geocode/json"
          locations=df['place'].to_list()
          performance_metric=df['performance value'].to_list()
          b=max(performance_metric)
          a=min(performance_metric)
          n=len(performance_metric)
          count=0
          y= []
          while count < n:
            z=(((performance_metric[count]-a)/(b-a))*(5-1))+1
            z=round(z)
            y.append(z)
            count = count +1
            
          # get latitudes and longitudes
          latitudes = []
          longitudes = []

          for location in locations:
              params = {'key': 'AIzaSyDyfSe7wZOsqlQaL-YY_-wzIEP0cH5i1lk', 'address':location}
              res = requests.get(url=url, params=params)
              location = res.json()['results'][0]['geometry']['location']
              latitudes.append(location['lat'])
              longitudes.append(location['lng'])

    

          gmap =gmplot.GoogleMapPlotter(statistics.mean(latitudes),statistics.mean(longitudes),5)
          gmap.apikey = "AIzaSyDyfSe7wZOsqlQaL-YY_-wzIEP0cH5i1lk"



        

          c=[1,2,3,4,5,6,7,8,9,10]

          count = 0

          dict = {}
          dict[1] = '#00FFFF'
          dict[2] = 'yellow'
          dict[3] = 'orange'
          dict[4] = '#6495ED'
          dict[5] = '#00FF00'


          while count < n:
              gmap.marker(latitudes[count],longitudes[count],title=locations[count],label=performance_metric[count],precision=2,color=dict[y[count]],info_window="performance value")
              count =count+1
          gmap.draw(r"templates\new2.html")

          return render(request,'new2.html',{'name':'uttej','status':'Data submitted successfully!'})

    #return render(request,'home.html',{'name':'uttej'})
    return render(request,'PVS.html',{'name':'uttej'})
    
#def add(request):

  #  val1 = int(request.POST["num1"])
  #  val2 = int(request.POST["num2"])
   # res =val1 + val2
    #return render(request,'result.html',{'result':res})

class DEA(object):

    def __init__(self, inputs, outputs):

        # supplied data
        self.inputs = inputs
        self.outputs = outputs

        # parameters
        self.n = inputs.shape[0]
        self.m = inputs.shape[1]
        self.r = outputs.shape[1]

        # iterators
        self.unit_ = range(self.n)
        self.input_ = range(self.m)
        self.output_ = range(self.r)

        # result arrays
        self.output_w = np.zeros((self.r, 1), dtype=np.float)  # output weights
        self.input_w = np.zeros((self.m, 1), dtype=np.float)  # input weights
        self.lambdas = np.zeros((self.n, 1), dtype=np.float)  # unit efficiencies
        self.efficiency = np.zeros_like(self.lambdas)  # thetas
    def __efficiency(self, unit):
        # compute efficiency
        denominator = np.dot(self.inputs, self.input_w)
        numerator = np.dot(self.outputs, self.output_w)

        return (numerator/denominator)[unit]
    def __target(self, x, unit):

        in_w, out_w, lambdas = x[:self.m], x[self.m:(self.m+self.r)], x[(self.m+self.r):]  # unroll the weights
        denominator = np.dot(self.inputs[unit], in_w)
        numerator = np.dot(self.outputs[unit], out_w)

        return numerator/denominator

    def __constraints(self, x, unit):

        in_w, out_w, lambdas = x[:self.m], x[self.m:(self.m+self.r)], x[(self.m+self.r):]  # unroll the weights
        constr = []  # init the constraint array

        # for each input, lambdas with inputs
        for input in self.input_:
            t = self.__target(x, unit)
            lhs = np.dot(self.inputs[:, input], lambdas)
            cons = t*self.inputs[unit, input] - lhs
            constr.append(cons)

        # for each output, lambdas with outputs
        for output in self.output_:
            lhs = np.dot(self.outputs[:, output], lambdas)
            cons = lhs - self.outputs[unit, output]
            constr.append(cons)

        # for each unit
        for u in self.unit_:
            constr.append(lambdas[u])

        return np.array(constr)

    def __optimize(self):

        d0 = self.m + self.r + self.n
        # iterate over units
        for unit in self.unit_:
            # weights
            x0 = np.random.rand(d0) - 0.5
            x0 = fmin_slsqp(self.__target, x0, f_ieqcons=self.__constraints, args=(unit,))
            # unroll weights
            self.input_w, self.output_w, self.lambdas = x0[:self.m], x0[self.m:(self.m+self.r)], x0[(self.m+self.r):]
            self.efficiency[unit] = self.__efficiency(unit)

    def fit(self):
        self.__optimize()  # optimize
        return self.efficiency

def home_efficiency(request):
    if request.method=="POST":
      excel_file = request.FILES["myfile"]
      check=check_extension(excel_file.name)
      if check is False:
          return render(request,'PEA.html',context={"error_in_file":"Allowed file extension is xlsx."})
      workbook = openpyxl.load_workbook(excel_file, data_only=True)
      worksheet=workbook.active
      excel_data = list()
      for row in worksheet.iter_rows():
          row_data = list()
          for cell in row:
              if cell.value=='':
                  cell.value=None
              row_data.append(cell.value)
          excel_data.append(row_data)
      upload_list=excel_data[0]
      valid_data=excel_data[1:]

      df = pd.DataFrame(columns=excel_data[0], data=excel_data[1:])
      print(df)
      upload_list=df.columns.tolist()
      check_list=['Port','IN1','IN2','IN3','OUT1','OUT2']
      if check_list!=upload_list:
          return render(request,'PEA.html',{'name':'uttej','error_in_file':'Headers not matched.Alllowed headers is "Port,IN1,IN2,IN3,OUT1,OUT2" !'})
      else:
          locations=df['Port'].to_list()
          print(locations)
          z=df.drop(df.loc[:,['Port']].head(0).columns, axis=1)
          print(z)
          inpt_df=df.iloc[:,1:4]
          X=np.array(inpt_df)
          print(X)
          outpt_df=df.iloc[:,4:]
          y=np.array(outpt_df)
          print(y)
          dea = DEA(X,y)
          e = dea.fit()
          print(e)
          k=e*100
          print('*********')
          print(k)
          kk=[]
          for val in k:
              if val[0]<0:
                  kk.append(40.68)
              elif val[0]>100:
                  kk.append(100)
              else:
                  kk.append(round(val[0],2))
          df['efficieny'] = np.around(kk)
          alldata=[]
          for i in range(df.shape[0]):
              tab = df.iloc[i]
              
              alldata.append(tab)
          context = {'data':alldata}
          print(alldata)
          url = "https://maps.googleapis.com/maps/api/geocode/json"
          locations=df['Port'].to_list()
          performance_metric=df['efficieny'].to_list()
          b=max(performance_metric)
          a=min(performance_metric)
          n=len(performance_metric)
          count=0
          u= []
          while count < n:
            z=(((performance_metric[count]-a)/(b-a))*(5-1))+1
            z=round(z)
            u.append(z)
            count = count +1
            
          # get latitudes and longitudes
          latitudes = []
          longitudes = []

          for location in locations:
              params = {'key': 'AIzaSyDyfSe7wZOsqlQaL-YY_-wzIEP0cH5i1lk', 'address':location}
              res = requests.get(url=url, params=params)
              location = res.json()['results'][0]['geometry']['location']
              latitudes.append(location['lat'])
              longitudes.append(location['lng'])

    

          gmap =gmplot.GoogleMapPlotter(statistics.mean(latitudes),statistics.mean(longitudes),5)
          gmap.apikey = "AIzaSyDyfSe7wZOsqlQaL-YY_-wzIEP0cH5i1lk"



        

          c=[1,2,3,4,5,6,7,8,9,10]

          count = 0

          dict = {}
          dict[1] = '#00FFFF'
          dict[2] = 'yellow'
          dict[3] = 'orange'
          dict[4] = '#6495ED'
          dict[5] = '#00FF00'
          remarks = {}
          remarks[1] = 'Efficiency: Worse'
          remarks[2] = 'Efficiency: To Improve'
          remarks[3] = 'Efficiency: Average'
          remarks[4] = 'Efficiency: Good'
          remarks[5] = 'Efficiency: Very Good'


          while count < n:
              gmap.marker(latitudes[count],longitudes[count],title=performance_metric[count],label=locations[count],precision=2,color=dict[u[count]],info_window=remarks[u[count]])
              count =count+1
          gmap.draw(r"templates\new3.html")

          return render(request,'taab.html',context)
    return render(request,'PEA.html',{'name':'uttej'})
